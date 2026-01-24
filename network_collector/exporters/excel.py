"""
Excel экспортер с форматированием.

Создаёт красиво оформленные Excel файлы с:
- Заголовками и стилями
- Автофильтром
- Цветовой индикацией (для статусов)
- Автоподбором ширины колонок

Пример использования:
    exporter = ExcelExporter(autofilter=True, freeze_header=True)
    exporter.export(data, "report.xlsx")
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .base import BaseExporter

logger = logging.getLogger(__name__)


# Предустановленные цвета
COLORS = {
    "header_bg": "4472C4",  # Синий фон заголовка
    "header_font": "FFFFFF",  # Белый текст заголовка
    "online": "C6EFCE",  # Зелёный (online)
    "offline": "FFC7CE",  # Красный (offline)
    "warning": "FFEB9C",  # Жёлтый (warning)
    "sticky": "C6EFCE",  # Зелёный (sticky)
    "dynamic": "FFFFFF",  # Белый (dynamic)
}


class ExcelExporter(BaseExporter):
    """
    Экспортер данных в Excel формат с форматированием.

    Создаёт профессионально оформленные таблицы с
    автофильтром, цветами и правильной шириной колонок.

    Attributes:
        autofilter: Включить автофильтр
        freeze_header: Закрепить строку заголовка
        auto_width: Автоподбор ширины колонок
        color_rules: Правила цветовой индикации

    Example:
        exporter = ExcelExporter()
        exporter.export(data, "report.xlsx")

        # С кастомными цветами
        exporter = ExcelExporter(
            color_rules={"status": {"online": "00FF00", "offline": "FF0000"}}
        )
    """

    file_extension = ".xlsx"

    def __init__(
        self,
        output_folder: str = "reports",
        autofilter: bool = True,
        freeze_header: bool = True,
        auto_width: bool = True,
        color_rules: Optional[Dict[str, Dict[str, str]]] = None,
    ):
        """
        Инициализация Excel экспортера.

        Args:
            output_folder: Папка для сохранения
            autofilter: Добавить автофильтр
            freeze_header: Закрепить первую строку
            auto_width: Автоподбор ширины колонок
            color_rules: Правила цветовой индикации
                        {column: {value: color_hex}}
        """
        super().__init__(output_folder, encoding="utf-8")

        self.autofilter = autofilter
        self.freeze_header = freeze_header
        self.auto_width = auto_width
        self.color_rules = color_rules or {}

        # Стили
        self._init_styles()

    def _init_styles(self) -> None:
        """Инициализирует стили для Excel."""
        # Стиль заголовка
        self.header_font = Font(bold=True, color=COLORS["header_font"])
        self.header_fill = PatternFill(
            start_color=COLORS["header_bg"],
            end_color=COLORS["header_bg"],
            fill_type="solid",
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Рамки
        thin_border = Side(style="thin", color="D9D9D9")
        self.cell_border = Border(
            left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
        )

    def _write(self, data: List[Dict[str, Any]], file_path: Path) -> None:
        """
        Записывает данные в Excel файл.

        Args:
            data: Данные для записи
            file_path: Путь к файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Получаем колонки
        columns = self._get_all_columns(data)

        # Заголовок
        self._write_header(ws, columns)

        # Данные
        self._write_data(ws, data, columns)

        # Форматирование
        if self.auto_width:
            self._adjust_column_widths(ws, columns, data)

        if self.autofilter:
            ws.auto_filter.ref = ws.dimensions

        if self.freeze_header:
            ws.freeze_panes = "A2"

        # Сохраняем
        wb.save(file_path)
        logger.debug(f"Excel записан: {len(data)} строк, {len(columns)} колонок")

    def _write_header(self, ws, columns: List[str]) -> None:
        """Записывает и форматирует заголовок."""
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column.upper())
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

    def _write_data(
        self,
        ws,
        data: List[Dict[str, Any]],
        columns: List[str],
    ) -> None:
        """Записывает данные с применением цветовых правил."""
        for row_idx, row in enumerate(data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row.get(column, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.cell_border

                # Применяем цветовые правила
                self._apply_color_rules(cell, column, value, row)

    def _apply_color_rules(
        self,
        cell,
        column: str,
        value: Any,
        row: Dict[str, Any],
    ) -> None:
        """
        Применяет цветовые правила к ячейке.

        Args:
            cell: Ячейка Excel
            column: Имя колонки
            value: Значение ячейки
            row: Вся строка данных
        """
        # Проверяем кастомные правила
        if column.lower() in self.color_rules:
            rules = self.color_rules[column.lower()]
            str_value = str(value).lower()
            if str_value in rules:
                color = rules[str_value]
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                return

        # Встроенные правила для status
        if column.lower() == "status":
            str_value = str(value).lower()
            if str_value == "online":
                cell.fill = PatternFill(
                    start_color=COLORS["online"],
                    end_color=COLORS["online"],
                    fill_type="solid",
                )
            elif str_value == "offline":
                cell.fill = PatternFill(
                    start_color=COLORS["offline"],
                    end_color=COLORS["offline"],
                    fill_type="solid",
                )

        # Встроенные правила для type (MAC)
        if column.lower() == "type":
            str_value = str(value).lower()
            if str_value == "sticky":
                # Зелёный только если online
                status = str(row.get("status", "")).lower()
                if status == "online":
                    cell.fill = PatternFill(
                        start_color=COLORS["sticky"],
                        end_color=COLORS["sticky"],
                        fill_type="solid",
                    )
                elif status == "offline":
                    cell.fill = PatternFill(
                        start_color=COLORS["offline"],
                        end_color=COLORS["offline"],
                        fill_type="solid",
                    )

    def _adjust_column_widths(
        self,
        ws,
        columns: List[str],
        data: List[Dict[str, Any]],
    ) -> None:
        """
        Автоподбор ширины колонок.

        Args:
            ws: Worksheet
            columns: Список колонок
            data: Данные
        """
        for col_idx, column in enumerate(columns, start=1):
            # Начинаем с длины заголовка
            max_length = len(column)

            # Проверяем данные
            for row in data:
                value = row.get(column, "")
                cell_length = len(str(value))
                if cell_length > max_length:
                    max_length = cell_length

            # Устанавливаем ширину (с небольшим запасом)
            adjusted_width = min(max_length + 2, 50)  # Максимум 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
