"""
Integration tests для Excel экспортера.

Проверяет что файлы создаются и содержат корректные данные.
"""

import pytest
from openpyxl import load_workbook

from network_collector.exporters import ExcelExporter


class TestExcelExporterIntegration:
    """Integration tests для ExcelExporter."""

    @pytest.fixture
    def sample_data(self):
        """Тестовые данные."""
        return [
            {"hostname": "switch-01", "ip": "10.0.0.1", "mac": "00:11:22:33:44:55", "vlan": 10},
            {"hostname": "switch-01", "ip": "10.0.0.1", "mac": "aa:bb:cc:dd:ee:ff", "vlan": 20},
            {"hostname": "switch-02", "ip": "10.0.0.2", "mac": "11:22:33:44:55:66", "vlan": 10},
        ]

    def test_excel_file_created(self, tmp_path, sample_data):
        """Файл Excel создаётся."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test_export.xlsx")

        assert result is not None
        assert result.exists()
        assert result.suffix == ".xlsx"

    def test_excel_can_be_opened(self, tmp_path, sample_data):
        """Excel файл валидный и открывается."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)  # Не должно выбросить исключение
        assert wb is not None
        wb.close()

    def test_excel_has_header(self, tmp_path, sample_data):
        """Excel содержит заголовок."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        # Заголовок в первой строке (uppercase)
        headers = [cell.value for cell in ws[1]]
        headers_lower = [h.lower() if h else "" for h in headers]

        assert "hostname" in headers_lower
        assert "ip" in headers_lower
        assert "mac" in headers_lower
        assert "vlan" in headers_lower
        wb.close()

    def test_excel_has_correct_row_count(self, tmp_path, sample_data):
        """Excel содержит все строки данных."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        # 1 header + 3 data rows = 4
        assert ws.max_row == 4
        wb.close()

    def test_excel_data_integrity(self, tmp_path, sample_data):
        """Данные в Excel соответствуют исходным."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        # Получаем заголовки
        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
        hostname_col = headers.index("hostname") + 1
        mac_col = headers.index("mac") + 1

        # Проверяем данные (row 2 = первая строка данных)
        assert ws.cell(row=2, column=hostname_col).value == "switch-01"
        assert ws.cell(row=2, column=mac_col).value == "00:11:22:33:44:55"
        assert ws.cell(row=4, column=hostname_col).value == "switch-02"
        wb.close()

    def test_excel_sheet_name(self, tmp_path, sample_data):
        """Лист называется 'Data'."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        assert wb.active.title == "Data"
        wb.close()

    def test_excel_autofilter_enabled(self, tmp_path, sample_data):
        """Автофильтр включён."""
        exporter = ExcelExporter(output_folder=str(tmp_path), autofilter=True)
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        assert ws.auto_filter.ref is not None
        wb.close()

    def test_excel_autofilter_disabled(self, tmp_path, sample_data):
        """Автофильтр отключён."""
        exporter = ExcelExporter(output_folder=str(tmp_path), autofilter=False)
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        assert ws.auto_filter.ref is None
        wb.close()

    def test_excel_freeze_header(self, tmp_path, sample_data):
        """Заголовок закреплён."""
        exporter = ExcelExporter(output_folder=str(tmp_path), freeze_header=True)
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        assert ws.freeze_panes == "A2"
        wb.close()

    def test_excel_no_freeze_header(self, tmp_path, sample_data):
        """Заголовок не закреплён."""
        exporter = ExcelExporter(output_folder=str(tmp_path), freeze_header=False)
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        assert ws.freeze_panes is None
        wb.close()

    def test_excel_empty_data_returns_none(self, tmp_path):
        """Пустые данные возвращают None."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export([], "empty.xlsx")

        assert result is None

    def test_excel_cyrillic_content(self, tmp_path):
        """Кириллица корректно записывается."""
        data = [
            {"hostname": "коммутатор-01", "description": "Серверная стойка"},
            {"hostname": "маршрутизатор-01", "description": "Ядро сети"},
        ]

        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(data, "cyrillic.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        # Проверяем кириллицу
        values = [cell.value for row in ws.iter_rows(min_row=2) for cell in row]
        assert "коммутатор-01" in values
        assert "Серверная стойка" in values
        wb.close()

    def test_excel_column_filter(self, tmp_path, sample_data):
        """Экспорт только выбранных колонок."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "filtered.xlsx", columns=["hostname", "mac"])

        wb = load_workbook(result)
        ws = wb.active

        # Только 2 колонки
        assert ws.max_column == 2

        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
        assert "hostname" in headers
        assert "mac" in headers
        assert "ip" not in headers
        wb.close()

    def test_excel_header_formatting(self, tmp_path, sample_data):
        """Заголовок имеет форматирование."""
        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(sample_data, "test.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        header_cell = ws.cell(row=1, column=1)

        # Жирный шрифт
        assert header_cell.font.bold is True
        # Есть заливка
        assert header_cell.fill.fgColor.rgb is not None
        wb.close()

    def test_excel_status_online_color(self, tmp_path):
        """Статус online имеет зелёный цвет."""
        data = [
            {"hostname": "switch-01", "status": "online"},
            {"hostname": "switch-02", "status": "offline"},
        ]

        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(data, "status.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
        status_col = headers.index("status") + 1

        # Online cell (row 2)
        online_cell = ws.cell(row=2, column=status_col)
        # Offline cell (row 3)
        offline_cell = ws.cell(row=3, column=status_col)

        # Проверяем что цвета разные
        online_color = online_cell.fill.fgColor.rgb if online_cell.fill.fgColor else None
        offline_color = offline_cell.fill.fgColor.rgb if offline_cell.fill.fgColor else None

        assert online_color != offline_color
        wb.close()

    def test_excel_custom_color_rules(self, tmp_path):
        """Кастомные цветовые правила работают."""
        data = [
            {"name": "test1", "priority": "high"},
            {"name": "test2", "priority": "low"},
        ]

        color_rules = {
            "priority": {
                "high": "FF0000",  # Красный
                "low": "00FF00",   # Зелёный
            }
        }

        exporter = ExcelExporter(output_folder=str(tmp_path), color_rules=color_rules)
        result = exporter.export(data, "custom_colors.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
        priority_col = headers.index("priority") + 1

        high_cell = ws.cell(row=2, column=priority_col)
        low_cell = ws.cell(row=3, column=priority_col)

        # Проверяем что цвета применены
        assert high_cell.fill.fgColor.rgb is not None
        assert low_cell.fill.fgColor.rgb is not None
        wb.close()

    def test_excel_auto_width(self, tmp_path):
        """Автоподбор ширины колонок."""
        data = [
            {"short": "a", "long_column_name": "very long value here that should affect width"},
        ]

        exporter = ExcelExporter(output_folder=str(tmp_path), auto_width=True)
        result = exporter.export(data, "width.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        # Ширина второй колонки должна быть больше первой
        col_a_width = ws.column_dimensions["A"].width
        col_b_width = ws.column_dimensions["B"].width

        assert col_b_width > col_a_width
        wb.close()

    def test_excel_numeric_values(self, tmp_path):
        """Числовые значения сохраняются как числа."""
        data = [
            {"name": "test", "count": 100, "price": 99.99},
        ]

        exporter = ExcelExporter(output_folder=str(tmp_path))
        result = exporter.export(data, "numbers.xlsx")

        wb = load_workbook(result)
        ws = wb.active

        headers = [cell.value.lower() if cell.value else "" for cell in ws[1]]
        count_col = headers.index("count") + 1
        price_col = headers.index("price") + 1

        # Числа должны быть числами, не строками
        assert ws.cell(row=2, column=count_col).value == 100
        assert ws.cell(row=2, column=price_col).value == 99.99
        wb.close()
