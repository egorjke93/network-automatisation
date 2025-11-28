"""
Модуль для сохранения MAC-адресов в Excel файл.
"""

import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .config import OUTPUT_FOLDER, SHOW_DEVICE_STATUS
from .utils import natural_sort_key


def save_to_excel(all_data: List[List], filename: str = None) -> str:
    """
    Сохраняет данные MAC-адресов в Excel файл с форматированием.
    
    Args:
        all_data: список записей [hostname, interface, description, mac, type, vlan, status?]
        filename: имя файла (если None - генерируется автоматически)
    
    Returns:
        Путь к сохранённому файлу
    """
    if not all_data:
        print("  ⚠️  Нет данных для сохранения")
        return None
    
    # Создаём папку для отчётов
    if OUTPUT_FOLDER:
        os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Генерируем имя файла
    if filename is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"mac_tables_{date_str}.xlsx"
    
    # Добавляем папку к пути
    if OUTPUT_FOLDER:
        filepath = os.path.join(OUTPUT_FOLDER, filename)
    else:
        filepath = filename
    
    # Создаём книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "MAC Addresses"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Заголовки
    headers = ["Device", "Interface", "Description", "MAC", "Type", "VLAN"]
    if SHOW_DEVICE_STATUS:
        headers.append("Status")
    ws.append(headers)
    
    # Применяем стили к заголовкам
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Сортировка: Device -> VLAN -> Interface (естественная)
    sorted_data = sorted(
        all_data,
        key=lambda x: (
            x[0],  # Device
            int(x[5]) if x[5].isdigit() else 0,  # VLAN как число
            natural_sort_key(x[1]),  # Interface
        ),
    )
    
    # Стили для выделения
    sticky_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    offline_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Индексы колонок
    type_col_idx = 4
    status_col_idx = 6 if SHOW_DEVICE_STATUS else None
    
    # Добавляем данные
    for row_idx, entry in enumerate(sorted_data, 2):
        ws.append(entry)
        
        # Определяем цвет
        entry_type = entry[type_col_idx] if len(entry) > type_col_idx else ""
        entry_status = (
            entry[status_col_idx]
            if status_col_idx and len(entry) > status_col_idx
            else ""
        )
        
        if SHOW_DEVICE_STATUS and entry_status == "offline":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = offline_fill
        elif entry_type == "sticky":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = sticky_fill
        
        # Добавляем границы
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # Ширина столбцов
    column_widths = {"A": 20, "B": 20, "C": 40, "D": 20, "E": 10, "F": 8}
    if SHOW_DEVICE_STATUS:
        column_widths["G"] = 10
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Закрепляем заголовок
    ws.freeze_panes = "A2"
    
    # Автофильтр
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(filepath)
    print(f"  ✅ Файл Excel сохранён: {filepath}")
    print(f"     Записей: {len(sorted_data)}")
    
    return filepath

